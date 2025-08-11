import os
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

# Set the root directory
root_dir = Path("E:/MHUData")
output_file = root_dir / "Combined_Workbook_with_Index.xlsx"

# Dictionary to store sheet data
sheets_data = {}

# Step 1: Collect data from each folder
for folder in root_dir.iterdir():
    if folder.is_dir():
        combined_df = pd.DataFrame()
        for file in folder.glob("*.xls*"):
            try:
                df = pd.read_excel(file)
                df['SourceFile'] = file.name
                combined_df = pd.concat([combined_df, df], ignore_index=True)
            except Exception as e:
                print(f"Error reading {file}: {e}")
        if not combined_df.empty:
            sheets_data[folder.name[:31]] = combined_df  # Excel sheet names limit

# Step 2: Create workbook and index sheet
wb = Workbook()
index_ws = wb.active
index_ws.title = "Home"

# Add hyperlinks to each sheet (start from row 1 now)
row = 1
for sheet_name in sheets_data.keys():
    cell = index_ws.cell(row=row, column=1)
    cell.value = sheet_name
    # Add quotes around sheet name for internal hyperlink
    cell.hyperlink = f"#'{sheet_name}'!A1"
    cell.font = Font(color="0000FF", underline="single")
    row += 1

# Step 3: Add each folder's data to a new sheet
for sheet_name, df in sheets_data.items():
    ws = wb.create_sheet(title=sheet_name)

    # Add "Back to Home" link
    ws['A1'] = "← Back to Home"
    ws['A1'].hyperlink = f"#'Home'!A1"
    ws['A1'].font = Font(color="0000FF", underline="single", bold=True)

    # Write DataFrame starting from row 3
    for col_num, col_name in enumerate(df.columns, start=1):
        ws.cell(row=3, column=col_num, value=col_name).font = Font(bold=True)
    for row_idx, row_data in enumerate(df.values, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

# Remove default sheet if still present
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Save the workbook
wb.save(output_file)
print(f"✅ Workbook created successfully: {output_file}")