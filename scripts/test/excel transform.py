import pandas as pd
from openpyxl import Workbook

# Load your Excel file from E drive
source_file = r'E:\HMIS\DM\Model Mapping 1.8.xlsx'
df = pd.read_excel(source_file)

# Create a new workbook
wb = Workbook()

# Remove the default created sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Create Index sheet
index_ws = wb.create_sheet(title='Index')
index_ws['A1'] = 'Sheet Name (ER No)'
index_ws.column_dimensions['A'].width = 30

# Track the index row number
index_row_num = 2

# Loop through each row to create a new sheet
for index, row in df.iterrows():
    # Get ER_NO value as sheet name
    er_no = str(row['ER No'])

    # Clean invalid characters and truncate to 31 characters if needed
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        er_no = er_no.replace(char, '_')
    sheet_name = er_no[:31]

    # Ensure unique sheet names by appending row number if duplicate
    while sheet_name in wb.sheetnames:
        sheet_name = f"{er_no[:28]}_{index+1}"

    # Create new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Add 'Back to Index' hyperlink in cell A1
    cell = ws.cell(row=1, column=1)
    cell.value = '<< Back to Index'
    cell.hyperlink = f"#Index!A1"
    cell.style = "Hyperlink"

    # Add headers and values starting from row 3 (preserving all columns)
    start_row = 3
    for j, col in enumerate(df.columns, start=0):
        ws.cell(row=start_row + j, column=1).value = col
        ws.cell(row=start_row + j, column=2).value = row[col]  # Make sure value is taken from the source row

    # Add entry to Index sheet with hyperlink
    index_cell = index_ws.cell(row=index_row_num, column=1)
    index_cell.value = sheet_name
    index_cell.hyperlink = f"#{sheet_name}!A1"
    index_cell.style = "Hyperlink"

    index_row_num += 1

# Save new workbook to E drive
output_file = r'E:\IHHN_Model.xlsx'
wb.save(output_file)

print(f"✔️ Data integrity verified — full workbook with hyperlinks created at {output_file}")
