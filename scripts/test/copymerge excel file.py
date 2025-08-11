import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import threading

# Helper function to ensure hyperlink style is applied correctly by openpyxl
# This avoids potential issues with style names not being registered by default
def apply_hyperlink_font_style(cell):
    cell.font = Font(color="0000FF", underline="single")


class ExcelCombinerApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("📁 Excel File Combiner 📊")
        self.geometry("700x500") # Set initial window size
        self.minsize(600, 400) # Minimum size for resizing

        # Configure grid layout for the main window
        self.grid_columnconfigure(0, weight=1)
        # Allocate weights to rows for proper resizing (results/status area)
        self.grid_rowconfigure(2, weight=1) # This row is for status/output display if needed, but here it acts as a spacer
        self.grid_rowconfigure(3, weight=0) # For status label
        self.grid_rowconfigure(4, weight=0) # For output path
        self.grid_rowconfigure(5, weight=0) # For "Created by MG" label

        # --- UI Elements ---

        # 1. Title Label
        self.title_label = ctk.CTkLabel(self, text="📁 Excel File Combiner 📊", font=ctk.CTkFont(size=24, weight="bold"), text_color="#004d40")
        self.title_label.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="n")

        # 2. Folder Selection Frame
        self.folder_frame = ctk.CTkFrame(self)
        self.folder_frame.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        self.folder_frame.grid_columnconfigure(1, weight=1) # Allow entry to expand

        self.folder_label = ctk.CTkLabel(self.folder_frame, text="Root Directory:", font=ctk.CTkFont(size=14, weight="bold"))
        self.folder_label.grid(row=0, column=0, padx=(10, 5), pady=10, sticky="w")

        self.selected_folder_path = ctk.StringVar()
        self.folder_entry = ctk.CTkEntry(self.folder_frame, textvariable=self.selected_folder_path, placeholder_text="Select folder containing Excel subfolders", font=ctk.CTkFont(size=12))
        self.folder_entry.grid(row=0, column=1, padx=5, pady=10, sticky="ew")

        self.browse_button = ctk.CTkButton(self.folder_frame, text="Browse Folder", command=self.browse_folder, font=ctk.CTkFont(size=12, weight="bold"), fg_color="#4CAF50", hover_color="#45a049")
        self.browse_button.grid(row=0, column=2, padx=(5, 10), pady=10)

        # 3. Process Button
        self.process_button = ctk.CTkButton(self, text="🚀 Process Excel Files", command=self.start_processing_thread, font=ctk.CTkFont(size=16, weight="bold"), fg_color="#2196F3", hover_color="#1976D2")
        self.process_button.grid(row=2, column=0, padx=20, pady=20, sticky="ew", ipadx=10, ipady=5)
        self.process_button.configure(state="disabled") # Disable initially

        # 4. Status Message
        self.status_message = ctk.StringVar(value="Ready.")
        self.status_label = ctk.CTkLabel(self, textvariable=self.status_message, font=ctk.CTkFont(size=12), text_color="gray", anchor="w")
        self.status_label.grid(row=3, column=0, padx=20, pady=(0, 5), sticky="ew")

        # 5. Output Path Display
        self.output_file_path = ctk.StringVar()
        self.output_path_display = ctk.CTkLabel(self, textvariable=self.output_file_path, font=ctk.CTkFont(size=12, slant="italic"), text_color="green", wraplength=650, justify="left", anchor="w")
        self.output_path_display.grid(row=4, column=0, padx=20, pady=(0, 10), sticky="ew")

        # 6. "Created by MG" Label
        self.creator_label = ctk.CTkLabel(self, text="Created by MG", font=ctk.CTkFont(size=10, slant="italic"), text_color="gray")
        self.creator_label.grid(row=5, column=0, padx=20, pady=(5, 10), sticky="s")

        self.processing_thread = None

    def browse_folder(self):
        """Opens a dialog to select a folder."""
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.selected_folder_path.set(folder_path)
            self.process_button.configure(state="normal") # Enable button once folder is selected
            self.status_message.set(f"Selected: {folder_path}")
            self.output_file_path.set("") # Clear previous output

    def start_processing_thread(self):
        """Starts the Excel combining process in a separate thread."""
        root_dir_str = self.selected_folder_path.get()

        if not root_dir_str:
            messagebox.showwarning("Input Error", "Please select a root directory first.")
            return

        if not os.path.isdir(root_dir_str):
            messagebox.showerror("Invalid Path", "The selected path is not a valid directory.")
            return

        # Disable button and update status
        self.process_button.configure(state="disabled", text="Processing...")
        self.browse_button.configure(state="disabled") # Disable browse during processing
        self.status_message.set("Processing... Please wait.")
        self.output_file_path.set("") # Clear previous output

        # Start the processing in a new thread
        self.processing_thread = threading.Thread(target=self._perform_excel_combination, args=(root_dir_str,))
        self.processing_thread.daemon = True # Allows thread to exit with main app
        self.processing_thread.start()

    def _perform_excel_combination(self, root_dir_str):
        """
        Performs the Excel combination operation. Runs in a separate thread.
        """
        root_dir = Path(root_dir_str)
        output_file_name = f"Combined_Workbook_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_file = root_dir / output_file_name

        try:
            sheets_data = {}
            found_excel_files_in_subfolders = False

            self.after(0, lambda: self.status_message.set("Scanning subfolders for Excel files..."))

            # Step 1: Collect data from each folder
            for folder in root_dir.iterdir():
                if folder.is_dir():
                    combined_df = pd.DataFrame()
                    subfolder_has_excel = False
                    for file in folder.glob("*.xls*"):
                        subfolder_has_excel = True
                        found_excel_files_in_subfolders = True
                        self.after(0, lambda f=file.name: self.status_message.set(f"Reading: {f} in {folder.name}..."))
                        try:
                            df = pd.read_excel(file)
                            df['SourceFile'] = file.name
                            combined_df = pd.concat([combined_df, df], ignore_index=True)
                        except Exception as e_info:
                            self.after(0, lambda info=e_info, fname=file.name: messagebox.showerror("File Read Error", f"Could not read {fname}: {info}\nPlease check if it's a valid Excel file and not open."))
                            self.after(0, self._re_enable_buttons)
                            return # Stop processing on file read error

                    if not subfolder_has_excel:
                        self.after(0, lambda fname=folder.name: self.status_message.set(f"No Excel files found in {fname}."))
                        self.update_idletasks() # Refresh UI

                    if not combined_df.empty:
                        # Excel sheet names limit to 31 characters
                        sheets_data[folder.name[:31]] = combined_df

            if not found_excel_files_in_subfolders:
                self.after(0, lambda: self.status_message.set("No Excel files found in the selected directory's subfolders."))
                self.after(0, lambda: messagebox.showinfo("No Files Found", "No Excel files were found in any subfolders of the selected directory."))
                self.after(0, self._re_enable_buttons)
                return

            if not sheets_data:
                self.after(0, lambda: self.status_message.set("No data could be extracted from the found Excel files."))
                self.after(0, lambda: messagebox.showinfo("No Data", "Excel files were found, but no data could be extracted to form sheets."))
                self.after(0, self._re_enable_buttons)
                return

            # Step 2: Create workbook and index sheet
            self.after(0, lambda: self.status_message.set("Creating workbook and index sheet..."))
            wb = Workbook()
            index_ws = wb.active
            index_ws.title = "Home"

            # Add hyperlinks to each sheet
            row_idx = 1
            for sheet_name in sheets_data.keys():
                cell = index_ws.cell(row=row_idx, column=1)
                cell.value = sheet_name
                cell.hyperlink = f"#'{sheet_name}'!A1"
                apply_hyperlink_font_style(cell)
                row_idx += 1

            # Step 3: Add each folder's data to a new sheet
            total_sheets_to_create = len(sheets_data)
            current_sheet_count = 0
            for sheet_name, df in sheets_data.items():
                current_sheet_count += 1
                self.after(0, lambda name=sheet_name, count=current_sheet_count, total=total_sheets_to_create: \
                    self.status_message.set(f"Adding sheet {count}/{total}: '{name}'..."))
                self.update_idletasks() # Ensure UI updates

                ws = wb.create_sheet(title=sheet_name)

                # Add "Back to Home" link
                back_to_home_cell = ws['A1']
                back_to_home_cell.value = "← Back to Home"
                back_to_home_cell.hyperlink = f"#'Home'!A1"
                back_to_home_cell.font = Font(color="0000FF", underline="single", bold=True)

                # Write DataFrame starting from row 3
                for col_num, col_name in enumerate(df.columns, start=1):
                    ws.cell(row=3, column=col_num, value=col_name).font = Font(bold=True)
                for row_idx, row_data in enumerate(df.values, start=4):
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

            # Remove default sheet if still present (if it was named 'Sheet')
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

            # Save the workbook
            self.after(0, lambda: self.status_message.set(f"Saving workbook to {output_file.name}..."))
            wb.save(output_file)

            self.after(0, lambda: self.status_message.set("✅ Workbook created successfully!"))
            self.after(0, lambda: self.output_file_path.set(f"Output saved to: {output_file.as_posix()}"))
            self.after(0, lambda: messagebox.showinfo("Success", f"Workbook created successfully at:\n{output_file.as_posix()}"))

        except Exception as e_info:
            self.after(0, lambda info=e_info: self.status_message.set(f"❌ An error occurred: {info}"))
            self.after(0, lambda info=e_info: messagebox.showerror("Error", f"An unexpected error occurred: {info}"))
        finally:
            self.after(0, self._re_enable_buttons)

    def _re_enable_buttons(self):
        """Re-enables UI elements after processing is complete or an error occurs."""
        self.process_button.configure(state="normal", text="🚀 Process Excel Files")
        self.browse_button.configure(state="normal") # Re-enable browse button

# Main execution block
if __name__ == "__main__":
    app = ExcelCombinerApp()
    app.mainloop()
